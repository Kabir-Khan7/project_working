"""
Bronze Writer — Ingests parsed files into the bronze_transactions table.

Part of the Medallion Architecture (Bronze -> Silver -> Gold) using DuckDB.
The Bronze Writer takes a parsed file and writes every row into bronze_transactions
exactly as received, preserving original column names and values.
"""

import csv
import io
import json
import os
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import duckdb

# Try to import openpyxl for Excel support
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Schema setup
# ---------------------------------------------------------------------------

BRONZE_TABLE_DDL = """
CREATE TABLE IF NOT EXISTS bronze_transactions (
    bronze_id           VARCHAR DEFAULT (CAST(uuid() AS VARCHAR)) PRIMARY KEY,
    source_type         VARCHAR NOT NULL,
    source_software     VARCHAR NOT NULL,
    source_file         VARCHAR NOT NULL,
    source_file_hash    VARCHAR NOT NULL,
    source_row_number   INTEGER,
    raw_content         VARCHAR NOT NULL,
    raw_headers         VARCHAR,
    ingestion_batch_id  VARCHAR NOT NULL,
    processing_status   VARCHAR NOT NULL DEFAULT 'pending',
    ingested_at         TIMESTAMP DEFAULT current_timestamp,
    updated_at          TIMESTAMP DEFAULT current_timestamp
);
"""

AUDIT_TABLE_DDL = """
CREATE TABLE IF NOT EXISTS ingestion_audit_log (
    audit_id            VARCHAR DEFAULT (CAST(uuid() AS VARCHAR)) PRIMARY KEY,
    batch_id            VARCHAR NOT NULL,
    source_file         VARCHAR NOT NULL,
    source_file_hash    VARCHAR NOT NULL,
    rows_written        INTEGER NOT NULL,
    status              VARCHAR NOT NULL,
    error_message       VARCHAR,
    started_at          TIMESTAMP NOT NULL,
    completed_at        TIMESTAMP DEFAULT current_timestamp
);
"""


def ensure_schema(conn: duckdb.DuckDBPyConnection):
    """Create tables if they don't exist."""
    conn.execute(BRONZE_TABLE_DDL)
    conn.execute(AUDIT_TABLE_DDL)


# ---------------------------------------------------------------------------
# ExcelParser
# ---------------------------------------------------------------------------

class ExcelParser:
    """Advanced Excel parser with title-row detection and multi-sheet support."""

    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel parsing. Install with: pip install openpyxl")

    def parse(self, file_path: str) -> list[dict]:
        """
        Parse an Excel file and return rows as list of dicts.

        - Picks the sheet with the most data rows.
        - Detects the header row (first row where >=50% cells are non-empty strings).
        - Skips blank rows.
        - Reads everything as strings.
        """
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Pick sheet with most data rows
        best_sheet = None
        best_row_count = -1
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            row_count = 0
            for row in ws.iter_rows():
                if any(cell.value is not None for cell in row):
                    row_count += 1
            if row_count > best_row_count:
                best_row_count = row_count
                best_sheet = sheet_name

        if best_sheet is None:
            wb.close()
            return []

        ws = wb[best_sheet]

        # Read all rows into memory (since read_only mode)
        all_rows = []
        for row in ws.iter_rows():
            all_rows.append([cell.value for cell in row])
        wb.close()

        if not all_rows:
            return []

        # Find header row: scan first 15 rows, find first where >=50% cells are strings
        header_idx = self._detect_header_row(all_rows[:15])
        if header_idx is None:
            # Fallback: use first row
            header_idx = 0

        headers = [str(cell) if cell is not None else f"column_{i}"
                   for i, cell in enumerate(all_rows[header_idx])]

        # Parse data rows
        rows = []
        for row_data in all_rows[header_idx + 1:]:
            # Skip blank rows
            if not any(cell is not None for cell in row_data):
                continue
            row_dict = {}
            for i, header in enumerate(headers):
                val = row_data[i] if i < len(row_data) else None
                row_dict[header] = str(val) if val is not None else ""
            rows.append(row_dict)

        return rows

    def _detect_header_row(self, rows: list[list]) -> Optional[int]:
        """Find the first row where >=50% of non-empty cells are strings."""
        for idx, row in enumerate(rows):
            non_empty = [cell for cell in row if cell is not None]
            if not non_empty:
                continue
            string_count = sum(1 for cell in non_empty if isinstance(cell, str))
            if string_count / len(non_empty) >= 0.5:
                return idx
        return None


# ---------------------------------------------------------------------------
# CSVParser
# ---------------------------------------------------------------------------

class CSVParser:
    """CSV parser with encoding detection and delimiter sniffing."""

    # Encoding fallback chain
    ENCODINGS = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]

    # Delimiters to try
    DELIMITERS = [",", "\t", ";", "|"]

    # Summary row indicators (lowercase)
    SUMMARY_KEYWORDS = {"total", "grand total", "sum", "kul", "majmua",
                        "کل", "مجموعہ"}

    def parse(self, file_path: str) -> list[dict]:
        """
        Parse a CSV file and return rows as list of dicts.

        - Tries multiple encodings.
        - Detects delimiter.
        - Skips summary/total rows at the bottom.
        - Handles BOM via utf-8-sig.
        """
        content = self._read_file(file_path)
        if content is None or not content.strip():
            return []

        delimiter = self._detect_delimiter(content)

        reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
        if not reader.fieldnames:
            return []

        rows = list(reader)

        # Remove trailing summary rows
        rows = self._strip_summary_rows(rows)

        return rows

    def _read_file(self, file_path: str) -> Optional[str]:
        """Try reading the file with multiple encodings."""
        for encoding in self.ENCODINGS:
            try:
                with open(file_path, "r", encoding=encoding) as f:
                    return f.read()
            except (UnicodeDecodeError, UnicodeError):
                continue
        return None

    def _detect_delimiter(self, content: str) -> str:
        """Detect the most likely delimiter using csv.Sniffer or frequency analysis."""
        # Try csv.Sniffer first
        try:
            sample = content[:8192]
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
            return dialect.delimiter
        except csv.Error:
            pass

        # Fallback: count occurrences in first line
        first_line = content.split("\n")[0]
        best_delim = ","
        best_count = 0
        for delim in self.DELIMITERS:
            count = first_line.count(delim)
            if count > best_count:
                best_count = count
                best_delim = delim
        return best_delim

    def _strip_summary_rows(self, rows: list[dict]) -> list[dict]:
        """Remove summary/total rows from the end of the data."""
        if not rows:
            return rows

        # Check last few rows from the bottom
        cut_index = len(rows)
        for i in range(len(rows) - 1, max(len(rows) - 6, -1), -1):
            row = rows[i]
            first_val = ""
            for val in row.values():
                if val and val.strip():
                    first_val = val.strip().lower()
                    break
            if first_val in self.SUMMARY_KEYWORDS:
                cut_index = i
            else:
                break

        return rows[:cut_index]


# ---------------------------------------------------------------------------
# BronzeWriter
# ---------------------------------------------------------------------------

class BronzeWriter:
    """Orchestrates file ingestion into the bronze_transactions table."""

    def __init__(self, conn: duckdb.DuckDBPyConnection):
        """
        Initialize BronzeWriter with a DuckDB connection.

        Args:
            conn: An open DuckDB connection.
        """
        self.conn = conn
        ensure_schema(self.conn)
        self._excel_parser = None
        self._csv_parser = CSVParser()

    @property
    def excel_parser(self) -> ExcelParser:
        if self._excel_parser is None:
            self._excel_parser = ExcelParser()
        return self._excel_parser

    def ingest(self, metadata: dict) -> dict:
        """
        Main entry point: parse a file and write all rows to bronze_transactions.

        Args:
            metadata: dict with keys:
                - file_path: str (absolute path to file)
                - file_name: str
                - source_type: str (e.g. "bank_statement", "ledger")
                - source_software: str (e.g. "Excel", "Tally")
                - file_hash: str (SHA-256 hash of file)
                - file_size_bytes: int

        Returns:
            dict with keys: success, rows_written, batch_id, error
        """
        file_path = metadata.get("file_path", "")
        file_name = metadata.get("file_name", "")
        source_type = metadata.get("source_type", "")
        source_software = metadata.get("source_software", "")
        file_hash = metadata.get("file_hash", "")
        batch_id = str(uuid.uuid4())
        started_at = datetime.now(timezone.utc)

        # Check for duplicate file
        if self._is_duplicate(file_hash):
            error_msg = f"Duplicate file detected: {file_name} (hash: {file_hash})"
            self._write_audit(batch_id, file_name, file_hash, 0, "rejected", error_msg, started_at)
            return {
                "success": False,
                "rows_written": 0,
                "batch_id": batch_id,
                "error": error_msg,
            }

        # Parse file
        try:
            rows = self._parse_file(file_path)
        except Exception as e:
            error_msg = f"Parse error: {str(e)}"
            self._write_audit(batch_id, file_name, file_hash, 0, "failed", error_msg, started_at)
            return {
                "success": False,
                "rows_written": 0,
                "batch_id": batch_id,
                "error": error_msg,
            }

        # Handle empty file
        if not rows:
            error_msg = "No data rows found in file"
            self._write_audit(batch_id, file_name, file_hash, 0, "failed", error_msg, started_at)
            return {
                "success": False,
                "rows_written": 0,
                "batch_id": batch_id,
                "error": error_msg,
            }

        # Get headers from first row
        headers = list(rows[0].keys())
        headers_json = json.dumps(headers, ensure_ascii=False)

        # Write rows to DuckDB
        rows_written = 0
        for idx, row in enumerate(rows, start=1):
            raw_content = json.dumps(row, ensure_ascii=False)
            self.conn.execute(
                """
                INSERT INTO bronze_transactions
                (bronze_id, source_type, source_software, source_file,
                 source_file_hash, source_row_number, raw_content, raw_headers,
                 ingestion_batch_id, processing_status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending')
                """,
                [
                    str(uuid.uuid4()),
                    source_type,
                    source_software,
                    file_name,
                    file_hash,
                    idx,
                    raw_content,
                    headers_json,
                    batch_id,
                ],
            )
            rows_written += 1

        # Write audit log
        self._write_audit(batch_id, file_name, file_hash, rows_written, "success", None, started_at)

        return {
            "success": True,
            "rows_written": rows_written,
            "batch_id": batch_id,
            "error": None,
        }

    def _parse_file(self, file_path: str) -> list[dict]:
        """Route to the correct parser based on file extension."""
        ext = Path(file_path).suffix.lower()
        if ext in (".xlsx", ".xls"):
            return self.excel_parser.parse(file_path)
        elif ext in (".csv", ".tsv", ".txt"):
            return self._csv_parser.parse(file_path)
        else:
            raise ValueError(f"Unsupported file type: {ext}")

    def _is_duplicate(self, file_hash: str) -> bool:
        """Check if a file with this hash has already been ingested."""
        result = self.conn.execute(
            "SELECT COUNT(*) FROM bronze_transactions WHERE source_file_hash = ?",
            [file_hash],
        ).fetchone()
        return result[0] > 0

    def _write_audit(self, batch_id: str, file_name: str, file_hash: str,
                     rows_written: int, status: str, error_msg: Optional[str],
                     started_at: datetime):
        """Write an entry to the ingestion audit log."""
        self.conn.execute(
            """
            INSERT INTO ingestion_audit_log
            (audit_id, batch_id, source_file, source_file_hash, rows_written,
             status, error_message, started_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                str(uuid.uuid4()),
                batch_id,
                file_name,
                file_hash,
                rows_written,
                status,
                error_msg,
                started_at,
            ],
        )
