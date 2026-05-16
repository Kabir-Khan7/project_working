"""
Tests for the Bronze Writer module.

Run with:
    python -m pytest test_bronze_writer.py -v -p no:asyncio
"""

import csv
import json
import os
import tempfile
import uuid

import duckdb
import pytest

import sys
from pathlib import Path

# Ensure direct import works when running tests from this directory
sys.path.insert(0, str(Path(__file__).parent))

from bronze_writer import BronzeWriter, CSVParser, ExcelParser, ensure_schema

# Check if openpyxl is available
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def db_conn():
    """Create an in-memory DuckDB connection with schema."""
    conn = duckdb.connect(":memory:")
    ensure_schema(conn)
    yield conn
    conn.close()


@pytest.fixture
def writer(db_conn):
    """Create a BronzeWriter instance."""
    return BronzeWriter(db_conn)


@pytest.fixture
def sample_csv(tmp_path):
    """Create a simple CSV file for testing."""
    file_path = tmp_path / "test.csv"
    file_path.write_text(
        "Date,Description,Amount\n"
        "2024-01-01,Payment received,5000\n"
        "2024-01-02,Office supplies,-200\n"
        "2024-01-03,Client payment,15000\n",
        encoding="utf-8",
    )
    return str(file_path)


@pytest.fixture
def sample_metadata(sample_csv):
    """Create metadata dict for a sample CSV file."""
    return {
        "file_path": sample_csv,
        "file_name": "test.csv",
        "source_type": "bank_statement",
        "source_software": "Excel",
        "file_hash": "abc123hash",
        "file_size_bytes": 150,
    }


# ---------------------------------------------------------------------------
# CSVParser Tests
# ---------------------------------------------------------------------------

class TestCSVParser:

    def test_basic_csv_parse(self, tmp_path):
        """Test parsing a simple CSV file."""
        f = tmp_path / "basic.csv"
        f.write_text("Name,Amount\nAlice,100\nBob,200\n", encoding="utf-8")

        parser = CSVParser()
        rows = parser.parse(str(f))

        assert len(rows) == 2
        assert rows[0]["Name"] == "Alice"
        assert rows[0]["Amount"] == "100"
        assert rows[1]["Name"] == "Bob"

    def test_csv_encoding_cp1252(self, tmp_path):
        """Test parsing a CSV with cp1252 encoding."""
        f = tmp_path / "cp1252.csv"
        content = "Name,City\nJohn,München\nAnna,Zürich\n"
        f.write_bytes(content.encode("cp1252"))

        parser = CSVParser()
        rows = parser.parse(str(f))

        assert len(rows) == 2
        assert rows[0]["City"] == "München"

    def test_csv_utf8_bom(self, tmp_path):
        """Test parsing a CSV with UTF-8 BOM."""
        f = tmp_path / "bom.csv"
        content = "Date,Amount\n2024-01-01,500\n"
        f.write_bytes(content.encode("utf-8-sig"))

        parser = CSVParser()
        rows = parser.parse(str(f))

        assert len(rows) == 1
        # BOM should be stripped, header should be clean
        assert "Date" in rows[0]

    def test_csv_tab_delimiter(self, tmp_path):
        """Test parsing a tab-delimited file."""
        f = tmp_path / "tabs.csv"
        f.write_text("Name\tAmount\tDate\nAlice\t100\t2024-01-01\n", encoding="utf-8")

        parser = CSVParser()
        rows = parser.parse(str(f))

        assert len(rows) == 1
        assert rows[0]["Name"] == "Alice"
        assert rows[0]["Amount"] == "100"

    def test_csv_strip_summary_rows(self, tmp_path):
        """Test that summary/total rows at the bottom are stripped."""
        f = tmp_path / "totals.csv"
        f.write_text(
            "Date,Description,Amount\n"
            "2024-01-01,Sale,1000\n"
            "2024-01-02,Sale,2000\n"
            "Total,,3000\n",
            encoding="utf-8",
        )

        parser = CSVParser()
        rows = parser.parse(str(f))

        assert len(rows) == 2
        assert all(r["Description"] != "" for r in rows)

    def test_csv_strip_urdu_summary(self, tmp_path):
        """Test that Urdu summary keywords are stripped."""
        f = tmp_path / "urdu_total.csv"
        f.write_text(
            "Date,Description,Amount\n"
            "2024-01-01,Sale,1000\n"
            "کل,,1000\n",
            encoding="utf-8",
        )

        parser = CSVParser()
        rows = parser.parse(str(f))

        assert len(rows) == 1

    def test_csv_empty_file(self, tmp_path):
        """Test parsing an empty CSV file."""
        f = tmp_path / "empty.csv"
        f.write_text("", encoding="utf-8")

        parser = CSVParser()
        rows = parser.parse(str(f))

        assert rows == []

    def test_csv_semicolon_delimiter(self, tmp_path):
        """Test parsing a semicolon-delimited file."""
        f = tmp_path / "semi.csv"
        f.write_text("Name;Amount;Date\nAlice;100;2024-01-01\n", encoding="utf-8")

        parser = CSVParser()
        rows = parser.parse(str(f))

        assert len(rows) == 1
        assert rows[0]["Name"] == "Alice"


# ---------------------------------------------------------------------------
# ExcelParser Tests
# ---------------------------------------------------------------------------

@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestExcelParser:

    def test_excel_basic_parse(self, tmp_path):
        """Test parsing a basic Excel file."""
        f = tmp_path / "basic.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Description", "Amount"])
        ws.append(["2024-01-01", "Payment", "5000"])
        ws.append(["2024-01-02", "Expense", "-200"])
        wb.save(str(f))

        parser = ExcelParser()
        rows = parser.parse(str(f))

        assert len(rows) == 2
        assert rows[0]["Date"] == "2024-01-01"
        assert rows[0]["Amount"] == "5000"

    def test_excel_title_row_detection(self, tmp_path):
        """Test that title rows (numeric/empty) are skipped to find real headers."""
        f = tmp_path / "titled.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        # Title/meta rows (numbers, not headers)
        ws.append([1, 2, 3])
        ws.append([None, None, None])
        ws.append(["Date", "Description", "Amount"])  # Real header at row 3
        ws.append(["2024-01-01", "Sale", "1000"])
        wb.save(str(f))

        parser = ExcelParser()
        rows = parser.parse(str(f))

        assert len(rows) == 1
        assert "Date" in rows[0]
        assert rows[0]["Date"] == "2024-01-01"

    def test_excel_skip_blank_rows(self, tmp_path):
        """Test that blank rows in data area are skipped."""
        f = tmp_path / "blanks.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Amount"])
        ws.append(["Alice", "100"])
        ws.append([None, None])  # blank row
        ws.append(["Bob", "200"])
        wb.save(str(f))

        parser = ExcelParser()
        rows = parser.parse(str(f))

        assert len(rows) == 2
        assert rows[0]["Name"] == "Alice"
        assert rows[1]["Name"] == "Bob"

    def test_excel_multi_sheet(self, tmp_path):
        """Test that the sheet with most data is selected."""
        f = tmp_path / "multi.xlsx"
        wb = openpyxl.Workbook()
        # First sheet: 1 data row
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.append(["Total", "100"])

        # Second sheet: 3 data rows (should be picked)
        ws2 = wb.create_sheet("Transactions")
        ws2.append(["Date", "Amount"])
        ws2.append(["2024-01-01", "500"])
        ws2.append(["2024-01-02", "600"])
        ws2.append(["2024-01-03", "700"])
        wb.save(str(f))

        parser = ExcelParser()
        rows = parser.parse(str(f))

        assert len(rows) == 3
        assert rows[0]["Date"] == "2024-01-01"


# ---------------------------------------------------------------------------
# BronzeWriter Tests
# ---------------------------------------------------------------------------

class TestBronzeWriter:

    def test_ingest_csv_success(self, writer, sample_metadata, db_conn):
        """Test successful CSV ingestion."""
        result = writer.ingest(sample_metadata)

        assert result["success"] is True
        assert result["rows_written"] == 3
        assert result["batch_id"] is not None
        assert result["error"] is None

    def test_raw_content_preserves_columns(self, writer, sample_metadata, db_conn):
        """Test that raw_content preserves original column names."""
        writer.ingest(sample_metadata)

        row = db_conn.execute(
            "SELECT raw_content FROM bronze_transactions LIMIT 1"
        ).fetchone()
        content = json.loads(row[0])

        assert "Date" in content
        assert "Description" in content
        assert "Amount" in content

    def test_raw_content_preserves_urdu(self, writer, db_conn, tmp_path):
        """Test that raw_content preserves Urdu column names."""
        f = tmp_path / "urdu.csv"
        f.write_text(
            "تاریخ,بیان,رقم\n"
            "2024-01-01,ادائیگی,5000\n",
            encoding="utf-8",
        )

        metadata = {
            "file_path": str(f),
            "file_name": "urdu.csv",
            "source_type": "ledger",
            "source_software": "Custom",
            "file_hash": "urdu_hash_123",
            "file_size_bytes": 100,
        }
        result = writer.ingest(metadata)
        assert result["success"] is True

        row = db_conn.execute(
            "SELECT raw_content FROM bronze_transactions WHERE source_file_hash = 'urdu_hash_123' LIMIT 1"
        ).fetchone()
        content = json.loads(row[0])

        # Verify Urdu headers are preserved
        assert "تاریخ" in content  # "taarikh" (date)
        assert "رقم" in content  # "raqam" (amount)

    def test_batch_id_groups_rows(self, writer, sample_metadata, db_conn):
        """Test that all rows from one file share the same batch_id."""
        result = writer.ingest(sample_metadata)
        batch_id = result["batch_id"]

        rows = db_conn.execute(
            "SELECT DISTINCT ingestion_batch_id FROM bronze_transactions WHERE source_file_hash = ?",
            [sample_metadata["file_hash"]],
        ).fetchall()

        assert len(rows) == 1
        assert rows[0][0] == batch_id

    def test_processing_status_pending(self, writer, sample_metadata, db_conn):
        """Test that processing_status defaults to 'pending'."""
        writer.ingest(sample_metadata)

        statuses = db_conn.execute(
            "SELECT DISTINCT processing_status FROM bronze_transactions"
        ).fetchall()

        assert len(statuses) == 1
        assert statuses[0][0] == "pending"

    def test_duplicate_file_rejection(self, writer, sample_metadata, db_conn):
        """Test that re-ingesting the same file hash is rejected."""
        # First ingest succeeds
        result1 = writer.ingest(sample_metadata)
        assert result1["success"] is True

        # Second ingest with same hash fails
        result2 = writer.ingest(sample_metadata)
        assert result2["success"] is False
        assert "Duplicate" in result2["error"]

    def test_empty_file_handling(self, writer, db_conn, tmp_path):
        """Test that empty files are handled gracefully."""
        f = tmp_path / "empty.csv"
        f.write_text("", encoding="utf-8")

        metadata = {
            "file_path": str(f),
            "file_name": "empty.csv",
            "source_type": "bank_statement",
            "source_software": "Excel",
            "file_hash": "empty_hash_456",
            "file_size_bytes": 0,
        }
        result = writer.ingest(metadata)

        assert result["success"] is False
        assert result["rows_written"] == 0
        assert "No data rows" in result["error"]

    def test_source_row_number_is_1_based(self, writer, sample_metadata, db_conn):
        """Test that source_row_number starts at 1."""
        writer.ingest(sample_metadata)

        rows = db_conn.execute(
            "SELECT source_row_number FROM bronze_transactions ORDER BY source_row_number"
        ).fetchall()

        assert rows[0][0] == 1
        assert rows[1][0] == 2
        assert rows[2][0] == 3

    def test_audit_log_entry_created(self, writer, sample_metadata, db_conn):
        """Test that an audit log entry is created on ingest."""
        result = writer.ingest(sample_metadata)

        audit = db_conn.execute(
            "SELECT batch_id, source_file, rows_written, status FROM ingestion_audit_log WHERE batch_id = ?",
            [result["batch_id"]],
        ).fetchone()

        assert audit is not None
        assert audit[0] == result["batch_id"]
        assert audit[1] == "test.csv"
        assert audit[2] == 3
        assert audit[3] == "success"

    def test_audit_log_on_failure(self, writer, db_conn, tmp_path):
        """Test that audit log records failures."""
        f = tmp_path / "empty.csv"
        f.write_text("", encoding="utf-8")

        metadata = {
            "file_path": str(f),
            "file_name": "empty.csv",
            "source_type": "bank_statement",
            "source_software": "Excel",
            "file_hash": "fail_hash_789",
            "file_size_bytes": 0,
        }
        result = writer.ingest(metadata)

        audit = db_conn.execute(
            "SELECT status, error_message FROM ingestion_audit_log WHERE batch_id = ?",
            [result["batch_id"]],
        ).fetchone()

        assert audit[0] == "failed"
        assert audit[1] is not None

    def test_unsupported_file_type(self, writer, db_conn, tmp_path):
        """Test that unsupported file types raise an error."""
        f = tmp_path / "data.pdf"
        f.write_text("fake pdf content", encoding="utf-8")

        metadata = {
            "file_path": str(f),
            "file_name": "data.pdf",
            "source_type": "invoice",
            "source_software": "Unknown",
            "file_hash": "pdf_hash_000",
            "file_size_bytes": 50,
        }
        result = writer.ingest(metadata)

        assert result["success"] is False
        assert "Unsupported file type" in result["error"]
